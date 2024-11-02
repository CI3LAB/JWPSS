import numpy as np
import pandas as pd

import openpyxl

workbook = openpyxl.Workbook()
worksheet = workbook.active

path="process_15_weight.npy"

data_dict = np.load(path, allow_pickle=True)


def min_col(data):
    min_row = np.argmin(data[:, 1])
    return min_row

# def read_data(data_dict,i):
#     objective = data_dict[i][:, 1]
#     reward = data_dict[i][:, 0]
#     makespan = data_dict[i][:, 2]
#     number_wp = data_dict[i][:, 3]
#     return objective


def read_data(data_dict,i):
    reward = data_dict[i][0]
    objective = data_dict[i][1]
    makespan = data_dict[i][2]
    number_wp = data_dict[i][3]
    return objective, reward, makespan, number_wp

for col in range(0, 1):
    mincol= min_col(data_dict[col])
    objective,reward,makespan,number_wp = read_data(data_dict[col],mincol)
    worksheet.cell(row=col+1, column=1, value=objective)
    worksheet.cell(row=col+1, column=2, value=reward)
    worksheet.cell(row=col+1, column=3, value=makespan)
    worksheet.cell(row=col+1, column=4, value=number_wp)
#
#
workbook.save("process.xlsx")


# 将数据写入到 .xlsx 文件,并保留所有数据

    # data = pd.DataFrame({"objective": objective})
    # data.to_excel("process_15.xlsx", sheet_name='Sheet1', startrow=0, startcol=i, header=False, index=False)





